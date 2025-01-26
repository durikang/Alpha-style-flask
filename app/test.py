import os
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from sklearn.linear_model import LinearRegression
import cx_Oracle
from datetime import datetime
from openpyxl.utils import get_column_letter
from flask import Flask, jsonify, request
import json
import folium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
import time
import logging
from selenium.webdriver.chrome.options import Options
from pathlib import Path
from openpyxl import Workbook, load_workbook


# ----------------------------
# Utility Functions
# ----------------------------
def create_output_paths(base_path="./analysis"):
    paths = {
        "output_dir": base_path,
        "output_dir_xlsx": os.path.join(base_path, "xlsx"),
        "output_dir_html": os.path.join(base_path, "html"),
        "output_dir_png": os.path.join(base_path, "png"),
    }
    for path in paths.values():
        os.makedirs(path, exist_ok=True)
    return paths


def save_excel(dataframe, path):
    """
    Save DataFrame to Excel with auto-adjusted column widths and print confirmation.
    """
    try:
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            dataframe.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for idx, column in enumerate(dataframe.columns, 1):
                max_length = dataframe[column].astype(str).map(len).max()
                adjusted_width = max(max_length, len(column)) + 2
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = adjusted_width
        print(f"엑셀 파일이 저장되었습니다: {path}")
    except Exception as e:
        print(f"엑셀 파일 저장 중 오류 발생: {e}")


def save_plotly_fig(fig, html_path, png_path, width=1800, height=1170):
    """
    Save Plotly figure as HTML and PNG files.
    """
    try:
        fig.write_html(html_path)
        print(f"HTML 파일이 저장되었습니다: {html_path}")
        try:
            fig.write_image(png_path, width=width, height=height)
            print(f"PNG 파일이 저장되었습니다: {png_path}")
        except Exception as e:
            print(f"PNG 파일 저장 중 오류 발생: {e}")
    except Exception as e:
        print(f"HTML 파일 저장 중 오류 발생: {e}")


# ----------------------------
# Oracle 데이터 조회 함수
# ----------------------------
def retrieve_oracle_data():
    """
    Oracle 데이터베이스에 연결하여 MEMBERS 및 ITEM 데이터를 조회합니다.
    """
    try:
        dsn = cx_Oracle.makedsn("localhost", 1521, service_name="xe")
        connection = cx_Oracle.connect(user="c##finalProject", password="1234", dsn=dsn)
        cursor = connection.cursor()

        # MEMBERS 데이터 조회
        oracle_query_1 = "SELECT BIRTH_DATE, USER_NO, ADDRESS, GENDER FROM MEMBERS"
        cursor.execute(oracle_query_1)
        columns_1 = [col[0] for col in cursor.description]
        data_1 = cursor.fetchall()
        oracle_data = pd.DataFrame(data_1, columns=columns_1)
        oracle_data.replace(['-'], np.nan, inplace=True)
        oracle_data.columns = ["생년월일", "유저번호", "지역", "성별"]

        # ITEM 및 SUB_CATEGORY 데이터 조회
        oracle_query_2 = """
            SELECT I.ITEM_NAME, I.SUB_CATEGORY_ID, SC.SUB_CATEGORY_NAME 
            FROM ITEM I 
            JOIN SUB_CATEGORY SC ON I.SUB_CATEGORY_ID = SC.ID
        """
        cursor.execute(oracle_query_2)
        columns_2 = [col[0] for col in cursor.description]
        data_2 = cursor.fetchall()
        oracle_item = pd.DataFrame(data_2, columns=columns_2)
        oracle_item.replace(['-'], np.nan, inplace=True)
        oracle_item.columns = ["품명", "SUB_ID", "카테고리"]
        oracle_query_3 = """

SELECT 
    FR.FIN_YEAR,
    M.ADDRESS,
    SUM(FR.SUPPLY_AMOUNT) AS TOTAL_SUPPLY_AMOUNT
FROM 
    FINANCIAL_RECORD FR
JOIN 
    MEMBERS M
ON 
    FR.USER_NO = M.USER_NO
WHERE 
    FR.TRANSACTION_TYPE = 1
GROUP BY 
    FR.FIN_YEAR, M.ADDRESS
ORDER BY 
    FR.FIN_YEAR, M.ADDRESS
                """
        cursor.execute(oracle_query_3)
        columns_3 = [col[0] for col in cursor.description]
        data_3 = cursor.fetchall()
        oracle_area = pd.DataFrame(data_3, columns=columns_3)
        oracle_area.replace(['-'], np.nan, inplace=True)
        oracle_area.columns = ["년도", "지역", "공급가액"]

        cursor.close()
        connection.close()

        # 나이 계산
        current_year = datetime.now().year
        oracle_data['생년월일'] = pd.to_datetime(oracle_data['생년월일'], errors='coerce')
        oracle_data['나이'] = current_year - oracle_data['생년월일'].dt.year
        oracle_data['나이'] = oracle_data['나이'].fillna(0).astype(int)

        return oracle_data, oracle_item, oracle_area

    except Exception as e:
        print(f"데이터베이스 연결 또는 데이터 조회 중 오류 발생: {e}")
        raise


# ----------------------------
# Financial Metrics Calculation
# ----------------------------
def calculate_sales(merged_data):
    sales_data = merged_data[merged_data['매입매출구분(1-매출/2-매입)'] == 1].copy()
    sales_data['년도'] = sales_data['년도'].astype(int)
    sales_data['매출'] = sales_data['수량'] * sales_data['단가']
    sales_by_year = sales_data.groupby('년도')['매출'].sum().reset_index()
    return sales_data, sales_by_year


def calculate_cost(merged_data):
    cost_data = merged_data[merged_data['매입매출구분(1-매출/2-매입)'] == 2].copy()
    cost_data['년도'] = cost_data['년도'].astype(int)
    cost_by_year = cost_data.groupby('년도')['공급가액'].sum().reset_index()
    cost_by_year.rename(columns={'공급가액': '판관비'}, inplace=True)
    return cost_by_year


def calculate_net_profit(sales_by_year, cost_by_year):
    net_profit = pd.merge(sales_by_year, cost_by_year, how='left', on='년도')
    net_profit['판관비'] = net_profit['판관비'].fillna(0).astype(float)
    net_profit['당기순이익'] = net_profit['매출'] - net_profit['판관비']
    return net_profit


# ----------------------------
# Plotting Functions
# ----------------------------
def plot_year_with_prediction(historical_year, historical_data, future_data, output_dir_html, output_dir_png):
    fig = go.Figure()

    fig.add_trace(go.Bar(
        x=['매출', '판관비', '당기순이익'],
        y=[
            historical_data['매출'] / 1e8,
            historical_data['판관비'] / 1e8,
            historical_data['당기순이익'] / 1e8
        ],
        name=f"{historical_year}년 (실제)",
        marker=dict(color=['red', 'blue', 'green'])
    ))

    fig.add_trace(go.Bar(
        x=['매출', '판관비', '당기순이익'],
        y=[
            (future_data['매출'] - historical_data['매출']) / 1e8,
            (future_data['판관비'] - historical_data['판관비']) / 1e8,
            (future_data['당기순이익'] - historical_data['당기순이익']) / 1e8
        ],
        name=f"{historical_year + 1}년 (예측)",
        marker=dict(color=['lightpink', 'lightblue', 'lightgreen'])
    ))

    fig.update_layout(
        title=f"{historical_year}년 vs {historical_year + 1}년 (스택예측)",
        xaxis_title="항목",
        yaxis_title="금액 (억 단위)",
        barmode='stack',
        font=dict(family="Arial, sans-serif", size=12),
        legend=dict(orientation="h", y=-0.2),
    )

    html_dir = os.path.join(output_dir_html, str(historical_year))
    png_dir = os.path.join(output_dir_png, str(historical_year))
    os.makedirs(html_dir, exist_ok=True)
    os.makedirs(png_dir, exist_ok=True)
    html_file = os.path.join(html_dir, f"{historical_year}_재무상태표.html")
    png_file = os.path.join(png_dir, f"{historical_year}_재무상태표.png")

    save_plotly_fig(fig, html_file, png_file)

def add_2025_predictions(net_profit):
    """
    net_profit 데이터프레임에 선형회귀를 사용해 2025년 예측 데이터를 추가합니다.
    """
    net_profit = net_profit.copy()
    net_profit['년도'] = net_profit['년도'].astype(int)

    # 숫자형 변환 및 NaN 처리
    for col in ['매출', '판관비', '당기순이익']:
        if col in net_profit.columns:
            net_profit[col] = pd.to_numeric(net_profit[col], errors='coerce').fillna(0)

    # 선형회귀 모델 생성
    lr = LinearRegression()

    # 새로운 데이터 저장용 리스트
    predictions = {'년도': [2025]}

    for col in ['매출', '판관비', '당기순이익']:
        if col in net_profit.columns:
            # X, y 설정
            X = net_profit[['년도']].values
            y = net_profit[col].values

            # y 값에 NaN이 없는지 확인
            if len(y) > 1 and not np.isnan(y).any():
                lr.fit(X, y)
                pred_2025 = lr.predict([[2025]])[0]
                predictions[col] = [pred_2025]

    # 새로운 예측 데이터를 DataFrame으로 추가
    predictions_df = pd.DataFrame(predictions)
    net_profit = pd.concat([net_profit, predictions_df], ignore_index=True)

    return net_profit


def plot_full_prediction_with_actuals(net_profit, output_dir_html, output_dir_png):
    """
    net_profit에 2025년 예측 데이터를 포함해 그래프를 그리고,
    모든 데이터(매출, 판관비, 당기순이익)에 대해 2024-2025는 점선으로 연결합니다.
    """
    try:
        # 2025년 예측 데이터 추가
        net_profit = add_2025_predictions(net_profit)

        # 억 단위 변환
        for col in ['매출', '판관비', '당기순이익']:
            if col in net_profit.columns:
                net_profit[col] = net_profit[col] / 1e8

        # 그래프 생성
        fig = go.Figure()

        # 실제 데이터: 2020~2024년
        for col, color, name in zip(['매출', '판관비', '당기순이익'],
                                    ['red', 'blue', 'green'],
                                    ['매출', '판관비', '당기순이익']):
            # 실제 데이터 (2020~2024)
            fig.add_trace(go.Scatter(
                x=net_profit.loc[net_profit['년도'] <= 2024, '년도'],
                y=net_profit.loc[net_profit['년도'] <= 2024, col],
                mode='lines+markers',
                name=f'{name}(실제)',
                line=dict(color=color)
            ))

            # 2024-2025 점선 연결
            if 2024 in net_profit['년도'].values and 2025 in net_profit['년도'].values:
                try:
                    x_values = [2024, 2025]
                    y_values = [
                        net_profit.loc[net_profit['년도'] == 2024, col].values[0],
                        net_profit.loc[net_profit['년도'] == 2025, col].values[0]
                    ]
                    fig.add_trace(go.Scatter(
                        x=x_values,
                        y=y_values,
                        mode='lines+markers',
                        name=f'{name}(예측)',
                        line=dict(dash='dot', color=color),
                        marker=dict(symbol='diamond', size=8, color=color)
                    ))
                except IndexError:
                    print(f"[WARN] 2025년 {name} 데이터가 누락되었습니다.")

        # 그래프 레이아웃 설정
        fig.update_layout(
            title="전체 재무데이터 (2025년 예측 추가)",
            xaxis_title="년도",
            yaxis_title="금액 (억)",
            hovermode='x unified',
            legend=dict(
                orientation="h",  # 가로 방향으로 설정
                yanchor="bottom",  # Y축 기준 하단 정렬
                y=-0.3,  # 그래프 하단으로 범례 위치 이동
                xanchor="center",  # X축 기준 중앙 정렬
                x=0.5  # 중앙에 배치
            )
        )

        # 저장
        html_file = os.path.join(output_dir_html, "연도별_재무상태표.html")
        png_file = os.path.join(output_dir_png, "연도별_재무상태표.png")
        fig.write_html(html_file)
        fig.write_image(png_file)

        print(f"HTML 파일이 저장되었습니다: {html_file}")
        print(f"PNG 파일이 저장되었습니다: {png_file}")

    except Exception as e:
        print(f"[ERROR] 그래프 생성 중 오류 발생: {e}")




# ----------------------------
# Gender Analysis Function
# ----------------------------
def analyze_gender(merged_data, oracle_data, output_dir_xlsx, output_dir_html, output_dir_png):
    """
    Perform gender-wise sales analysis, perform linear regression prediction,
    and generate corresponding plots and save data.
    """
    # Filter sales data
    sales_administrative = merged_data[merged_data['매입매출구분(1-매출/2-매입)'] == 1].copy()

    # Merge with oracle_data to get user information
    merged_gender = pd.merge(sales_administrative, oracle_data, on='유저번호')

    # Ensure '년도' is numeric
    merged_gender['년도'] = pd.to_numeric(merged_gender['년도'], errors='coerce')

    # Aggregate sales by year and gender
    year_gender_spending = merged_gender.groupby(['년도', '성별'])['공급가액'].sum().reset_index()

    # Initialize '예측 공급가액' column with NaN
    year_gender_spending['예측 공급가액'] = np.nan

    # ----------------------------
    # **수정된 부분 시작** #
    # ----------------------------
    # 각 년도별로 다음 년도의 예측값을 현재 년도의 '예측 공급가액'에 저장
    # 성별별 데이터 예측
    for gender in year_gender_spending['성별'].unique():
        gender_data = year_gender_spending[year_gender_spending['성별'] == gender].copy()
        gender_data.sort_values('년도', inplace=True)
        years = gender_data['년도'].unique()
        for y in years:
            # 학습 데이터: 현재 년도까지
            train_df = gender_data[gender_data['년도'] <= y]
            if len(train_df) < 2:
                # 데이터가 부족하면 예측값을 실제값으로 설정
                if not train_df.empty:
                    pred_value = train_df['공급가액'].iloc[-1]
                else:
                    pred_value = 0.0
            else:
                X = train_df[['년도']].astype(int).values
                y_val = train_df['공급가액'].astype(float).values
                lr = LinearRegression()
                lr.fit(X, y_val)
                future_year = y + 1
                pred_value = lr.predict([[future_year]])[0]

            # 현재 년도의 '예측 공급가액'에 다음 년도의 예측값 저장
            year_gender_spending.loc[
                (year_gender_spending['년도'] == y) & (year_gender_spending['성별'] == gender),
                '예측 공급가액'
            ] = pred_value

    # Ensure '공급가액' and '예측 공급가액' are float
    year_gender_spending['공급가액'] = year_gender_spending['공급가액'].astype(float).fillna(0)
    year_gender_spending['예측 공급가액'] = year_gender_spending['예측 공급가액'].astype(float).fillna(0)

    # 성별별 합산 데이터 추가
    total_by_gender = (
        year_gender_spending.groupby('성별')[['공급가액', '예측 공급가액']]
        .sum()
        .reset_index()
    )
    total_by_gender['년도'] = '전체'  # "년도"를 "전체"로 설정

    # 컬럼 순서 맞추기
    total_by_gender = total_by_gender[['년도', '성별', '공급가액', '예측 공급가액']]

    # 설명 컬럼 추가
    description = "\n".join(
        f"전체 {row['성별']} - 공급가액: {row['공급가액']}, 예측 공급가액: {row['예측 공급가액']}"
        for _, row in total_by_gender.iterrows()
    )

    # "설명" 컬럼 생성: 첫 번째 행에만 설명 추가
    total_by_gender['설명'] = ""
    total_by_gender.loc[0, '설명'] = description

    # 데이터 Excel로 저장
    gender_output = os.path.join(output_dir_xlsx, "성별별_판매량.xlsx")
    total_by_gender.to_excel(gender_output, index=False)
    print(f"성별별 전체 매출 데이터 Excel 파일 저장 완료: {gender_output}")

    # Generate and save pie charts and Excel files for each year including predictions
    for year in sorted(year_gender_spending['년도'].dropna().unique()):
        year = int(year)
        year_dir_html = os.path.join(output_dir_html, str(year))
        year_dir_png = os.path.join(output_dir_png, str(year))
        year_dir_xlsx = os.path.join(output_dir_xlsx, str(year))  # Directory for Excel
        os.makedirs(year_dir_html, exist_ok=True)
        os.makedirs(year_dir_png, exist_ok=True)
        os.makedirs(year_dir_xlsx, exist_ok=True)  # Ensure Excel directory exists
        print(year_dir_html, year_dir_png, year_dir_xlsx)

        # 현재 루프에서 year_data를 정의
        year_data = year_gender_spending[year_gender_spending['년도'] == year].copy()

        # 설명 컬럼 추가: 성별 데이터를 한 줄로 요약
        description = "\n".join(
            f"{row['년도']}년도 {row['성별']} - 공급가액: {row['공급가액']}, 예측 공급가액: {row['예측 공급가액']}"
            for _, row in year_data.iterrows()
        )

        # 설명 컬럼 생성: 첫 번째 행에만 설명 추가
        year_data['설명'] = ""
        year_data.loc[year_data.index[0], '설명'] = description

        # Save to Excel (actual and predicted values)
        year_excel_output = os.path.join(year_dir_xlsx, f"{year}_성별_매출.xlsx")
        try:
            save_excel(year_data, year_excel_output)
            print(f"{year}년 성별 매출 데이터 Excel 파일 저장 완료: {year_excel_output}")
        except Exception as e:
            print(f"Excel 파일 저장 중 오류 발생: {e}")

        # Generate and save pie chart (using actual values)
        try:
            fig = go.Figure(data=[
                go.Pie(
                    labels=year_data['성별'],
                    values=year_data['공급가액'],
                    hole=0.3,
                    textinfo='label+percent'
                )
            ])
            fig.update_layout(
                title=f"{year}년 성별 매출 비중",
                font=dict(family="Arial, sans-serif", size=12),
                legend=dict(
                    x=0,
                    y=1,
                    xanchor="left",
                    yanchor="top"
                )
            )

            html_file = os.path.join(year_dir_html, f"{year}_성별_매출.html")
            png_file = os.path.join(year_dir_png, f"{year}_성별_매출.png")
            save_plotly_fig(fig, html_file, png_file)
            print(f"{year}년 성별 매출 그래프 저장 완료: {html_file}, {png_file}")
        except Exception as e:
            print(f"그래프 생성 중 오류 발생: {e}")

    # Generate and save line chart for gender and predicted
    # Generate and save line chart for gender and predicted
    try:
        # 선형 회귀를 사용해 2025년 데이터 예측
        for gender in year_gender_spending['성별'].unique():
            gender_data = year_gender_spending[year_gender_spending['성별'] == gender]
            if len(gender_data) >= 2:
                X = gender_data[['년도']].astype(int).values
                y = gender_data['공급가액'].astype(float).values
                lr = LinearRegression()
                lr.fit(X, y)
                pred_2025 = lr.predict([[2025]])[0]
                year_gender_spending = pd.concat([
                    year_gender_spending,
                    pd.DataFrame({'년도': [2025], '성별': [gender], '공급가액': [None], '예측 공급가액': [pred_2025]})
                ], ignore_index=True)

        # Pivot actual '공급가액' and predicted '예측 공급가액'
        gender_pivot = year_gender_spending.pivot_table(
            index='년도',
            columns='성별',
            values=['공급가액', '예측 공급가액'],
            aggfunc='sum'
        ).fillna(0)

        # Flatten the multi-level columns
        gender_pivot.columns = [f"{val}_{col}" for val, col in gender_pivot.columns]

        # Create separate DataFrames for actual and predicted
        gender_actual = gender_pivot.filter(like='공급가액').copy()
        gender_predicted = gender_pivot.filter(like='예측 공급가액').copy()

        # Convert to 억 단위
        gender_actual_plot = gender_actual / 1e8
        gender_predicted_plot = gender_predicted / 1e8

        # Plot actual and predicted
        fig = go.Figure()
        colors = {gender: color for gender, color in
                  zip(year_gender_spending['성별'].unique(), ['blue', 'red'])}
        for gender in year_gender_spending['성별'].unique():
            actual_col = f"공급가액_{gender}"
            predicted_col = f"예측 공급가액_{gender}"

            # Add actual data (2020~2024)
            if actual_col in gender_actual_plot.columns:
                fig.add_trace(go.Scatter(
                    x=gender_actual_plot.index[gender_actual_plot.index < 2025].astype(int),
                    y=gender_actual_plot.loc[gender_actual_plot.index < 2025, actual_col],
                    mode='lines+markers',
                    name=f'{gender} 매출 (실제)',
                    line=dict(color=colors.get(gender, 'black'))
                ))

            # Add predicted data (2025 as dotted line)
            if predicted_col in gender_predicted_plot.columns and 2025 in gender_predicted_plot.index:
                fig.add_trace(go.Scatter(
                    x=[2024, 2025],
                    y=[
                        gender_actual_plot.loc[2024, actual_col] if 2024 in gender_actual_plot.index else None,
                        gender_predicted_plot.loc[2025, predicted_col]
                    ],
                    mode='lines+markers',
                    name=f'{gender} 매출 (예측)',
                    line=dict(dash='dot', color=colors.get(gender, 'black')),
                    marker=dict(symbol='diamond', size=8, color=colors.get(gender, 'black'))
                ))

        fig.update_layout(
            title='연도별 성별 매출 (실제 + 2025 예측)',
            xaxis_title='년도',
            yaxis_title='금액 (억 단위)',
            font=dict(family="Arial, sans-serif", size=12),
            legend=dict(orientation="h", y=-0.2),
        )

        # Save updated line chart
        html_file = os.path.join(output_dir_html, "연도별_성별_매출.html")
        png_file = os.path.join(output_dir_png, "연도별_성별_매출.png")
        save_plotly_fig(fig, html_file, png_file)
        print(f"성별 매출 라인 차트 저장 완료 (2025 예측 포함): {html_file}, {png_file}")

    except Exception as e:
        print(f"[ERROR] 예측 그래프 생성 중 오류 발생: {e}")


# ----------------------------
# Age Group Analysis Function
# ----------------------------
def analyze_age_group(merged_data, oracle_data, output_dir_xlsx, output_dir_html, output_dir_png):
    """
    Perform age-group-wise sales analysis, perform linear regression prediction,
    and generate corresponding plots and save data.
    """
    try:
        # 1. 매출 데이터 필터링
        sales_administrative = merged_data[merged_data['매입매출구분(1-매출/2-매입)'] == 1].copy()

        # 2. 유저 정보와 병합
        merged_age = pd.merge(sales_administrative, oracle_data, on='유저번호', how='left')

        # 3. '년도'를 숫자형으로 변환
        merged_age['년도'] = pd.to_numeric(merged_age['년도'], errors='coerce')

        # 4. 나이대 정의
        bins = [10, 20, 30, 40, 50]
        labels = ['10대', '20대', '30대', '40대']
        merged_age['나이대'] = pd.cut(merged_age['나이'], bins=bins, labels=labels, right=False, include_lowest=True)

        # 5. 연도 및 나이대별 공급가액 집계
        year_age_spending = merged_age.groupby(['년도', '나이대'])['공급가액'].sum().reset_index()

        # 6. NaN '나이대' 제거
        year_age_spending = year_age_spending.dropna(subset=['나이대'])

        # 7. '예측 공급가액' 초기화
        year_age_spending['예측 공급가액'] = np.nan

        # 8. 예측 수행
        for age_group in year_age_spending['나이대'].unique():
            try:
                if pd.isna(age_group):
                    continue
                age_data = year_age_spending[year_age_spending['나이대'] == age_group].copy()
                age_data.sort_values('년도', inplace=True)
                years = age_data['년도'].unique()
                for y in years:
                    try:
                        # 학습 데이터: 현재 년도까지
                        train_df = age_data[age_data['년도'] <= y]
                        if len(train_df) < 2:
                            # 데이터가 부족하면 예측값을 실제값으로 설정
                            if not train_df.empty:
                                pred_value = train_df['공급가액'].iloc[-1]
                            else:
                                pred_value = 0.0
                        else:
                            X = train_df[['년도']].astype(int).values
                            y_val = train_df['공급가액'].astype(float).values
                            lr = LinearRegression()
                            lr.fit(X, y_val)
                            future_year = y + 1
                            pred_value = lr.predict([[future_year]])[0]

                        # 예측값 저장
                        year_age_spending.loc[
                            (year_age_spending['년도'] == y) & (year_age_spending['나이대'] == age_group),
                            '예측 공급가액'
                        ] = pred_value
                    except Exception as inner_e:
                        print(f"년도 {y}에 대한 예측 중 오류 발생: {inner_e}")
            except Exception as outer_e:
                print(f"나이대 {age_group} 처리 중 오류 발생: {outer_e}")

        # 9. 데이터 타입 변환
        year_age_spending['공급가액'] = year_age_spending['공급가액'].astype(float).fillna(0)
        year_age_spending['예측 공급가액'] = year_age_spending['예측 공급가액'].astype(float).fillna(0)

        # 나이대별 합산 데이터 추가
        total_by_age_group = (
            year_age_spending.groupby('나이대')[['공급가액', '예측 공급가액']]
            .sum()
            .reset_index()
        )
        total_by_age_group['년도'] = '전체'  # "년도"를 "전체"로 설정

        # 컬럼 순서 맞추기
        total_by_age_group = total_by_age_group[['년도', '나이대', '공급가액', '예측 공급가액']]

        # 설명 컬럼 추가
        description = "\n".join(
            f"전체 {row['나이대']} - 공급가액: {row['공급가액']}, 예측 공급가액: {row['예측 공급가액']}"
            for _, row in total_by_age_group.iterrows()
        )

        # "설명" 컬럼 생성: 첫 번째 행에만 설명 추가
        total_by_age_group['설명'] = ""
        total_by_age_group.loc[0, '설명'] = description

        # 데이터 Excel로 저장
        age_output = os.path.join(output_dir_xlsx, "나이대별_판매량.xlsx")
        total_by_age_group.to_excel(age_output, index=False)
        print(f"연령대별 전체 매출 데이터 Excel 파일 저장 완료: {age_output}")

        # 11. 연도별 파이 차트 및 Excel 저장
        for year in sorted(year_age_spending['년도'].dropna().unique()):
            try:
                year = int(year)  # 현재 처리 중인 연도를 정수로 변환
                year_dir_html = os.path.join(output_dir_html, str(year))
                year_dir_png = os.path.join(output_dir_png, str(year))
                year_dir_xlsx = os.path.join(output_dir_xlsx, str(year))
                os.makedirs(year_dir_html, exist_ok=True)
                os.makedirs(year_dir_png, exist_ok=True)
                os.makedirs(year_dir_xlsx, exist_ok=True)
                print(f"Directories 생성 완료: {year_dir_html}, {year_dir_png}, {year_dir_xlsx}")

                # 현재 연도 데이터 추출
                current_year_data = year_age_spending[year_age_spending['년도'] == year].copy()

                # 설명 컬럼 생성: 모든 데이터를 하나로 결합
                description = "\n".join(
                    f"{row['년도']}년도 {row['나이대']} - 공급가액: {row['공급가액']}, 예측 공급가액: {row['예측 공급가액']}"
                    for _, row in current_year_data.iterrows()
                )

                # 첫 번째 행에만 설명 추가, 나머지 행은 빈 값으로 설정
                current_year_data['설명'] = ""
                current_year_data.loc[current_year_data.index[0], '설명'] = description

                # Excel 저장
                year_excel_output = os.path.join(year_dir_xlsx, f"{year}_나이대별_판매량.xlsx")
                save_excel(current_year_data, year_excel_output)
                print(f"{year}년 연령대별 매출 데이터 Excel 파일 저장 완료: {year_excel_output}")

                # 원형 그래프 생성 및 저장
                try:
                    fig = go.Figure(data=[
                        go.Pie(
                            labels=current_year_data['나이대'],
                            values=current_year_data['공급가액'],
                            hole=0.3,
                            textinfo='label+percent'
                        )
                    ])
                    fig.update_layout(
                        title=f"{year}년 연령대별 매출 비중",
                        font=dict(family="Arial, sans-serif", size=12),
                        legend=dict(
                            x=0,
                            y=1,
                            xanchor="left",
                            yanchor="top"
                        )
                    )

                    html_file = os.path.join(year_dir_html, f"{year}_나이대별_매출.html")
                    png_file = os.path.join(year_dir_png, f"{year}_나이대별_매출.png")
                    save_plotly_fig(fig, html_file, png_file)
                    print(f"{year}년 연령대별 원형 그래프 저장 완료: {html_file}, {png_file}")
                except Exception as e:
                    print(f"그래프 생성 중 오류 발생: {e}")
            except Exception as e:
                print(f"년도 {year} 처리 중 오류 발생: {e}")
    except Exception as e:
        print(f"그래프 생성 중 오류 발생: {e}")

    try:
        # 선형 회귀를 사용해 2025년 데이터 예측
        for age_group in year_age_spending['나이대'].unique():
            age_data = year_age_spending[year_age_spending['나이대'] == age_group]
            if len(age_data) >= 2:
                X = age_data[['년도']].astype(int).values
                y = age_data['공급가액'].astype(float).values
                lr = LinearRegression()
                lr.fit(X, y)
                pred_2025 = lr.predict([[2025]])[0]
                year_age_spending = pd.concat([
                    year_age_spending,
                    pd.DataFrame({'년도': [2025], '나이대': [age_group], '공급가액': [None], '예측 공급가액': [pred_2025]})
                ], ignore_index=True)

        # Pivot actual '공급가액' and predicted '예측 공급가액'
        age_pivot = year_age_spending.pivot_table(
            index='년도',
            columns='나이대',
            values=['공급가액', '예측 공급가액'],
            aggfunc='sum'
        ).fillna(0)

        # Flatten the multi-level columns
        age_pivot.columns = [f"{val}_{col}" for val, col in age_pivot.columns]

        # Create separate DataFrames for actual and predicted
        age_actual = age_pivot.filter(like='공급가액').copy()
        age_predicted = age_pivot.filter(like='예측 공급가액').copy()

        # Convert to 억 단위
        age_actual_plot = age_actual / 1e8
        age_predicted_plot = age_predicted / 1e8

        # Plot actual and predicted
        fig = go.Figure()
        colors = {age_group: color for age_group, color in
                  zip(year_age_spending['나이대'].unique(), ['blue', 'red', 'green', 'yellow', 'purple', 'orange'])}

        for age_group in year_age_spending['나이대'].unique():
            actual_col = f"공급가액_{age_group}"
            predicted_col = f"예측 공급가액_{age_group}"

            # Add actual data (2020~2024)
            if actual_col in age_actual_plot.columns:
                fig.add_trace(go.Scatter(
                    x=age_actual_plot.index[age_actual_plot.index < 2025].astype(int),
                    y=age_actual_plot.loc[age_actual_plot.index < 2025, actual_col],
                    mode='lines+markers',
                    name=f'{age_group} 매출 (실제)',
                    line=dict(color=colors.get(age_group, 'black'))
                ))

            # Add predicted data (2025 as dotted line)
            if predicted_col in age_predicted_plot.columns and 2025 in age_predicted_plot.index:
                fig.add_trace(go.Scatter(
                    x=[2024, 2025],
                    y=[
                        age_actual_plot.loc[2024, actual_col] if 2024 in age_actual_plot.index else None,
                        age_predicted_plot.loc[2025, predicted_col]
                    ],
                    mode='lines+markers',
                    name=f'{age_group} 매출 (예측)',
                    line=dict(dash='dot', color=colors.get(age_group, 'black')),
                    marker=dict(symbol='diamond', size=8, color=colors.get(age_group, 'black'))
                ))

        fig.update_layout(
            title='연도별 연령대별 매출 (실제 + 2025 예측)',
            xaxis_title='년도',
            yaxis_title='금액 (억 단위)',
            font=dict(family="Arial, sans-serif", size=12),
            legend=dict(orientation="h", y=-0.2),
        )

        # Save line chart
        html_file = os.path.join(output_dir_html, "연도별_나이대별_매출.html")
        png_file = os.path.join(output_dir_png, "연도별_나이대별_매출.png")
        save_plotly_fig(fig, html_file, png_file)
        print(f"연령대별 매출 라인 차트 저장 완료 (2025 예측 포함): {html_file}, {png_file}")

    except Exception as e:
        print(f"[ERROR] 예측 그래프 생성 중 오류 발생: {e}")


# ----------------------------
# Category Analysis Function
# ----------------------------
def analyze_category(merged_data, oracle_item, output_dir_xlsx, output_dir_html, output_dir_png):
    """
    Perform category-wise sales analysis, perform linear regression prediction,
    and generate corresponding plots and save data.
    """
    try:
        # (1) 매출 데이터 추출
        sales_data = merged_data[merged_data['매입매출구분(1-매출/2-매입)'] == 1].copy()
        print("매출 데이터 추출 완료.")

        min_year_cat = sales_data['년도'].min()
        max_year_cat = sales_data['년도'].max()
        print(f"카테고리별 분석을 위한 연도 범위: {min_year_cat} ~ {max_year_cat}")

        # 카테고리별 (년도, 품목→카테고리, 공급가액) 집계
        oracle_item.columns = ["품명", "SUB_ID", "카테고리"]  # 컬럼 이름 재정의
        merged_cat_df = (
            sales_data
            .groupby(['년도', '품명'])['공급가액']
            .sum()
            .reset_index()
            .merge(oracle_item[['품명', '카테고리']], on='품명', how='left')
        )
        merged_cat_df['카테고리'] = merged_cat_df['카테고리'].fillna('미분류')
        print("카테고리별 집계 완료.")

        # 년도+카테고리별 집계
        merged_cat_df = merged_cat_df.groupby(['년도', '카테고리'])['공급가액'].sum().reset_index()
        print("년도 및 카테고리별 공급가액 집계 완료.")

        # 전체 연도 범위
        all_years = sorted(merged_cat_df['년도'].unique())  # 예: [2020, 2021, 2022, 2023, 2024]
        print(f"전체 연도: {all_years}")
        final_list = []

        # 2) 카테고리별 (올해 → 내년) 예측
        cat_list = merged_cat_df['카테고리'].unique()
        print(f"분석할 카테고리 목록: {cat_list}")

        for cat in cat_list:
            print(f"카테고리 '{cat}' 분석 시작.")
            cat_data = merged_cat_df[merged_cat_df['카테고리'] == cat].copy()
            cat_data.sort_values('년도', inplace=True)
            print(f"카테고리 '{cat}' 데이터:\n{cat_data}")

            # 선형 회귀로 카테고리별 데이터를 학습
            if len(cat_data) < 2:
                print(f"카테고리 '{cat}'의 데이터가 부족하여 예측을 수행하지 않습니다.")
                for y in all_years:
                    row_ = cat_data[cat_data['년도'] == y]
                    if len(row_) > 0:
                        real_val = row_['공급가액'].values[0]
                    else:
                        real_val = 0.0
                    # 올해 행에 내년 예측 필요 → +1년
                    next_pred = real_val  # 또는 0
                    final_list.append({
                        '카테고리': cat,
                        '년도': y,
                        '실제공급가액': real_val,
                        '예측공급가액': next_pred
                    })
                continue

            # 데이터가 2개 이상인 경우 선형 회귀를 통해 내년 예측
            X = cat_data[['년도']].values
            y_val = cat_data['공급가액'].values
            lr = LinearRegression()
            lr.fit(X, y_val)
            print(f"카테고리 '{cat}'에 대한 선형 회귀 모델 학습 완료.")

            for y in all_years:
                # 실제 값
                row_ = cat_data[cat_data['년도'] == y]
                if len(row_) > 0:
                    real_val = row_['공급가액'].values[0]
                else:
                    real_val = 0.0

                # 예측 값 (y + 1년에 대한 예측)
                next_pred = lr.predict([[y + 1]])[0]

                final_list.append({
                    '카테고리': cat,
                    '년도': y,
                    '실제공급가액': real_val,
                    '예측공급가액': next_pred
                })
            print(f"카테고리 '{cat}'에 대한 예측 완료.")

        cat_df_final = pd.DataFrame(final_list)
        print("카테고리별 예측 데이터프레임 생성 완료.")

        # wide pivot 생성
        pivot_actual = cat_df_final.pivot_table(
            index='년도',
            columns='카테고리',
            values='실제공급가액',
            aggfunc='first'
        ).fillna(0)

        pivot_pred = cat_df_final.pivot_table(
            index='년도',
            columns='카테고리',
            values='예측공급가액',
            aggfunc='first'
        ).fillna(0)
        pivot_pred.columns = ['예측' + c for c in pivot_pred.columns]
        print("실제 및 예측 데이터 pivot 완료.")

        # 실제와 예측 데이터 병합
        cat_wide = pd.concat([pivot_actual, pivot_pred], axis=1)
        cat_wide.reset_index(inplace=True)
        print("실제 및 예측 데이터 병합 완료.")

        # 컬럼 정렬
        actual_cats = pivot_actual.columns.tolist()
        pred_cats = pivot_pred.columns.tolist()
        final_cols = ['년도'] + actual_cats + pred_cats
        print(f"최종 컬럼 목록: {final_cols}")

        # 실제 존재하는지 체크
        final_cols = [c for c in final_cols if c in cat_wide.columns]
        cat_wide = cat_wide[final_cols]
        print("데이터프레임 컬럼 정렬 완료.")

        # --------- 1. 년도별 카테고리별 예측 Excel 파일 저장 ---------
        print("년도별 카테고리별 예측 Excel 파일 저장 시작.")
        for year in all_years:
            year = int(year)
            year_dir_html = os.path.join(output_dir_html, str(year))
            year_dir_png = os.path.join(output_dir_png, str(year))
            year_dir_xlsx = os.path.join(output_dir_xlsx, str(year))  # Directory for Excel
            os.makedirs(year_dir_html, exist_ok=True)
            os.makedirs(year_dir_png, exist_ok=True)
            os.makedirs(year_dir_xlsx, exist_ok=True)  # Ensure Excel directory exists

            year_df = cat_wide[cat_wide['년도'] == year].copy()
            if year_df.empty:
                print(f"{year}년에 대한 데이터가 없어 Excel 파일을 생성하지 않습니다.")
                continue

            # Melt the DataFrame to have '카테고리', '실제공급가액', '예측공급가액'
            melted_actual = year_df.melt(id_vars=['년도'], value_vars=actual_cats,
                                         var_name='카테고리', value_name='실제공급가액')
            melted_pred = year_df.melt(id_vars=['년도'], value_vars=pred_cats,
                                       var_name='카테고리', value_name='예측공급가액')
            melted_pred['카테고리'] = melted_pred['카테고리'].str.replace('예측', '')
            combined = pd.merge(melted_actual, melted_pred, on=['년도', '카테고리'], how='left')
            combined['예측공급가액'] = combined['예측공급가액'].fillna(0)

            # 설명 컬럼 추가
            description = "\n".join(
                f"{row['년도']}년도 {row['카테고리']} - 실제공급가액: {row['실제공급가액']}, 예측공급가액: {row['예측공급가액']}"
                for _, row in combined.iterrows()
            )

            # "설명" 컬럼 생성: 첫 번째 행에만 설명 추가
            combined['설명'] = ""
            combined.loc[combined.index[0], '설명'] = description

            # Debugging: 출력해보기
            print(f"{year}년 데이터 확인:\n{combined}")

            # Save to Excel
            excel_filename = f"{year}_카테고리별_판매량.xlsx"  # 파일 이름에 '가로형' 추가
            excel_path = os.path.join(year_dir_xlsx, excel_filename)
            save_excel(combined, excel_path)

            # --------- 1-1. 년도별 카테고리별 그래프 생성 ---------
            try:
                fig_year = go.Figure()

                # 실제 공급가액
                fig_year.add_trace(go.Bar(
                    x=combined['카테고리'],
                    y=combined['실제공급가액'] / 1e8,  # 억 단위로 변환
                    name='실제 공급가액',
                    marker=dict(color='blue')
                ))

                # 예측 공급가액
                fig_year.add_trace(go.Bar(
                    x=combined['카테고리'],
                    y=combined['예측공급가액'] / 1e8,  # 억 단위로 변환
                    name='예측 공급가액',
                    marker=dict(color='lightblue')
                ))

                fig_year.update_layout(
                    title=f"{year}년 카테고리별 공급가액 (실제 + 예측)",
                    xaxis_title="카테고리",
                    yaxis_title="공급가액 (억)",
                    barmode='group',
                    font=dict(family="Arial, sans-serif", size=12),
                    legend=dict(orientation="h", y=-0.2),
                )

                # 그래프 저장
                fig_year_html = os.path.join(year_dir_html, f"{year}_카테고리별_판매량.html")
                fig_year_png = os.path.join(year_dir_png, f"{year}_카테고리별_판매량.png")
                save_plotly_fig(fig_year, fig_year_html, fig_year_png)
                print(f"{year}년 카테고리별 판매량 그래프 저장 완료: {fig_year_html}, {fig_year_png}")

            except Exception as e:
                print(f"{year}년 카테고리별 그래프 생성 중 오류 발생: {e}")

        print("전체 카테고리별 합산 데이터 생성 및 저장 시작.")
        # 모든 연도의 데이터를 합산
        sum_cat = cat_df_final.groupby('카테고리').agg({
            '실제공급가액': 'sum',
            '예측공급가액': 'sum'
        }).reset_index()
        print(f"합산된 카테고리별 데이터:\n{sum_cat}")

        # 합산 데이터를 '카테고리별_판매량_예측.xlsx' 파일로 저장
        sum_cat['년도'] = "전체"  # 모든 데이터의 "년도" 값을 "전체"로 설정

        # 설명 컬럼 추가
        description = "\n".join(
            f"전체 {row['카테고리']} - 실제공급가액: {row['실제공급가액']}, 예측공급가액: {row['예측공급가액']}"
            for _, row in sum_cat.iterrows()
        )

        # "설명" 컬럼 생성: 첫 번째 행에만 설명 추가
        sum_cat['설명'] = ""
        if not sum_cat.empty:
            sum_cat.loc[0, '설명'] = description

        # 컬럼 순서 재배치 (년도가 맨 앞으로 오도록)
        sum_cat = sum_cat[['년도', '카테고리', '실제공급가액', '예측공급가액', '설명']]

        # 합산 데이터를 '카테고리별_판매량_예측.xlsx' 파일로 저장
        sum_excel_path = os.path.join(output_dir_xlsx, "연도별_카테고리별_판매량.xlsx")
        save_excel(sum_cat, sum_excel_path)
        print(f"카테고리별 전체 합산 데이터 Excel 파일 저장 완료: {sum_excel_path}")

        # --------- 3. 전체 카테고리별 합산 그래프 생성 ---------
        print("전체 카테고리별 합산 그래프 생성 시작.")
        fig_sum = go.Figure()

        fig_sum.add_trace(go.Bar(
            x=sum_cat['카테고리'],
            y=sum_cat['실제공급가액'] / 1e8,
            name='실제 공급가액',
            marker=dict(color='blue')
        ))

        fig_sum.add_trace(go.Bar(
            x=sum_cat['카테고리'],
            y=sum_cat['예측공급가액'] / 1e8,
            name='예측 공급가액',
            marker=dict(color='lightblue')
        ))

        fig_sum.update_layout(
            title="카테고리별 전체 공급가액 예측",
            xaxis_title="카테고리",
            yaxis_title="공급가액 (억)",
            barmode='group',
            font=dict(family="Arial, sans-serif", size=12),
            legend=dict(orientation="h", y=-0.2),
        )

        # 합산 그래프 저장
        sum_html_path = os.path.join(output_dir_html, "연도별_카테고리별_판매량.html")
        sum_png_path = os.path.join(output_dir_png, "연도별_카테고리별_판매량.png")
        save_plotly_fig(fig_sum, sum_html_path, sum_png_path)
        print(f"카테고리별 전체 합산 그래프 저장 완료: {sum_html_path}, {sum_png_path}")

    except Exception as e:
        print(f"카테고리별 분석 중 오류 발생: {e}")


# ----------------------------
# VIP Users Analysis Function
# ----------------------------
def analyze_vip_users(merged_data, oracle_data, output_dir_xlsx, output_dir_html, output_dir_png):
    """
    1) 연도별로 VIP 분석 (상위 30% 실제 vs 예측):
       - 누적분포(파랑) vs 예측분포(노랑)
       - 10%,20%,30% 점선 (빨강:실제, 주황:예측)
       - 엑셀(10%,20%,30% 지점 실제/예측)

    2) 전체(모든 연도 합산)에 대해서도 같은 방식으로
       - user_spending_total (유저별 총액)
       - 상위 30% 실제 vs 예측
       - 그래프(파랑 vs 노랑) + 10%,20%,30% 점선
       - 엑셀 저장
    """
    try:
        # 1) 매출 데이터만 추출
        sales_data = merged_data[merged_data['매입매출구분(1-매출/2-매입)'] == 1].copy()

        # 2) 유저 정보와 병합
        merged_vip = pd.merge(sales_data, oracle_data, on='유저번호', how='inner')

        # 3) '년도' 숫자형 변환
        merged_vip['년도'] = pd.to_numeric(merged_vip['년도'], errors='coerce')

        # 4) 퍼센트 지점 (10%,20%,30%)
        percentages = [0.1, 0.2, 0.3]

        # ----------------------------
        # (A) 연도별 분석
        # ----------------------------
        years = sorted(merged_vip['년도'].dropna().unique())
        for year in years:
            year = int(year)
            year_data = merged_vip[merged_vip['년도'] == year]

            # 디렉토리 생성
            year_dir_html = os.path.join(output_dir_html, str(year))
            year_dir_png = os.path.join(output_dir_png, str(year))
            year_dir_xlsx = os.path.join(output_dir_xlsx, str(year))

            os.makedirs(year_dir_html, exist_ok=True)
            os.makedirs(year_dir_png, exist_ok=True)
            os.makedirs(year_dir_xlsx, exist_ok=True)

            # 유저별 소비금액 내림차순
            user_spending = (
                year_data
                .groupby('유저번호')['공급가액']
                .sum()
                .sort_values(ascending=False)
                .reset_index()
            )
            user_spending['누적금액'] = user_spending['공급가액'].cumsum()
            actual_vals = user_spending['누적금액'] / 1e8  # 억 단위
            x_vals = np.linspace(0, 1, len(user_spending))

            # (A-1) 상위 30% 실제 & 예측
            cutoff_30 = int(np.ceil(len(user_spending) * 0.3))
            actual_30 = 0.0
            predicted_30 = 0.0
            if cutoff_30 > 0:
                top_30 = user_spending.iloc[:cutoff_30]
                actual_30 = top_30['공급가액'].sum()

                # 선형회귀 -> 다음 연도 예측
                top_30_user_list = top_30['유저번호'].tolist()
                hist_30_data = merged_vip[merged_vip['유저번호'].isin(top_30_user_list)]
                yearly_30 = hist_30_data.groupby('년도')['공급가액'].sum().reset_index()
                if len(yearly_30) >= 2:
                    lr = LinearRegression()
                    X = yearly_30[['년도']].astype(int).values
                    y_val_lr = yearly_30['공급가액'].astype(float).values
                    lr.fit(X, y_val_lr)
                    future_year = year + 1
                    predicted_30 = lr.predict([[future_year]])[0]
                else:
                    if not yearly_30.empty:
                        predicted_30 = yearly_30['공급가액'].iloc[-1]
                    else:
                        predicted_30 = 0.0

            # (A-2) 스케일링 비율
            ratio = 1.0
            if actual_30 > 0:
                ratio = predicted_30 / actual_30

            predicted_vals = ratio * actual_vals  # 노랑 영역

            # (A-3) 엑셀 저장 (10%,20%,30%)
            excel_rows = []
            for p in percentages:
                cutoff_idx = int(np.ceil(len(user_spending) * p))
                if cutoff_idx <= 0:
                    continue
                actual_v = actual_vals[cutoff_idx - 1]
                pred_v = predicted_vals[cutoff_idx - 1]
                excel_rows.append({
                    "비율": f"{int(p * 100)}%",
                    "실제공급가액": actual_v,  # 컬럼 이름 수정 (억 제거)
                    "예측공급가액": pred_v  # 컬럼 이름 수정 (억 제거)
                })

            df_excel = pd.DataFrame(excel_rows)

            # 억 단위에서 원 단위로 값 변환
            df_excel['실제공급가액'] *= 1e8
            df_excel['예측공급가액'] *= 1e8

            # 설명 컬럼 추가
            description = "\n".join(
                f"{row['비율']} - 실제공급가액: {row['실제공급가액']}, 예측공급가액: {row['예측공급가액']}"
                for _, row in df_excel.iterrows()
            )
            df_excel['설명'] = ""
            df_excel.loc[0, '설명'] = description  # 첫 번째 행에만 설명 추가

            # 엑셀 파일 저장
            excel_path = os.path.join(year_dir_xlsx, f"{year}_VIP_유저.xlsx")
            df_excel_copy_file = df_excel.copy()
            df_excel_copy_file.insert(0, '년도', year)
            save_excel(df_excel_copy_file, excel_path)
            print(f"{year}년 VIP 유저 예측 데이터 Excel 파일 저장 완료: {excel_path}")

            # (A-4) 그래프
            try:
                fig = go.Figure()

                # 파랑 영역 (실제)
                fig.add_trace(go.Scatter(
                    x=x_vals,
                    y=actual_vals,
                    fill='tozeroy',
                    mode='none',
                    fillcolor='blue',
                    name='실제 누적분포'
                ))
                # 노랑 영역 (예측)
                fig.add_trace(go.Scatter(
                    x=x_vals,
                    y=predicted_vals,
                    fill='tonexty',
                    mode='none',
                    fillcolor='yellow',
                    name='예측 누적분포'
                ))

                # 점선: 10%,20%,30% => 실제=빨강, 예측=주황
                for p in percentages:
                    cutoff_idx = int(np.ceil(len(user_spending) * p))
                    if cutoff_idx <= 0:
                        continue
                    cutoff_x = cutoff_idx / len(user_spending)

                    # 실제(빨강)
                    ya = actual_vals[cutoff_idx - 1]
                    fig.add_trace(go.Scatter(
                        x=[cutoff_x, cutoff_x],
                        y=[0, ya],
                        mode='lines',
                        line=dict(color='red', dash='dash'),
                        name=f"{int(p * 100)}% 경계(실제)"
                    ))
                    # 예측(주황)
                    yp = predicted_vals[cutoff_idx - 1]
                    fig.add_trace(go.Scatter(
                        x=[cutoff_x, cutoff_x],
                        y=[0, yp],
                        mode='lines',
                        line=dict(color='orange', dash='dash'),
                        name=f"{int(p * 100)}% 경계(예측)"
                    ))

                fig.update_layout(
                    title=f"{year}년 VIP 유저 (실제 vs 예측) 누적분포",
                    xaxis=dict(title="유저 비율", range=[0, 1]),
                    yaxis=dict(title="누적 금액 (억)", range=[0, max(predicted_vals) * 1.05]),
                    font=dict(size=12),
                    legend=dict(orientation="h", y=-0.2),
                    margin=dict(l=50, r=50, t=50, b=100)
                )

                # 그래프 저장
                html_path = os.path.join(year_dir_html, f"{year}_VIP_유저.html")
                png_path = os.path.join(year_dir_png, f"{year}_VIP_유저.png")
                save_plotly_fig(fig, html_path, png_path)
            except Exception as e:
                print(f"그래프 생성 중 오류 발생: {e}")

        # ----------------------------
        # (B) 전체(모든 연도 합산) 분석
        # ----------------------------
        try:
            # 1) 유저별 총합
            user_spending_total = (
                merged_vip
                .groupby('유저번호')['공급가액']
                .sum()
                .sort_values(ascending=False)
                .reset_index()
            )
            user_spending_total['누적금액'] = user_spending_total['공급가액'].cumsum()
            actual_vals_total = user_spending_total['누적금액'] / 1e8
            x_vals_total = np.linspace(0, 1, len(user_spending_total))

            # (B-1) 상위 30% 실제 & 예측
            cutoff_30_tot = int(np.ceil(len(user_spending_total) * 0.3))
            actual_30_tot = 0.0
            predicted_30_tot = 0.0
            if cutoff_30_tot > 0:
                top_30_total = user_spending_total.iloc[:cutoff_30_tot]
                actual_30_tot = top_30_total['공급가액'].sum()

                # 모든 연도 데이터 중 상위 30% 유저들의 연도별 합 → 선형회귀
                top30_users_list = top_30_total['유저번호'].tolist()
                hist_30_total = merged_vip[merged_vip['유저번호'].isin(top30_users_list)]
                yearly_30_total = hist_30_total.groupby('년도')['공급가액'].sum().reset_index()

                if len(yearly_30_total) >= 2:
                    lr_all = LinearRegression()
                    X_ = yearly_30_total[['년도']].astype(int).values
                    y_ = yearly_30_total['공급가액'].astype(float).values
                    lr_all.fit(X_, y_)
                    future_year_ = max(yearly_30_total['년도'].astype(int).unique()) + 1
                    predicted_30_tot = lr_all.predict([[future_year_]])[0]
                else:
                    if not yearly_30_total.empty:
                        predicted_30_tot = yearly_30_total['공급가액'].iloc[-1]
                    else:
                        predicted_30_tot = 0.0

            # (B-2) 스케일링 비율
            ratio_all = 1.0
            if actual_30_tot > 0:
                ratio_all = predicted_30_tot / actual_30_tot

            predicted_vals_total = ratio_all * actual_vals_total

            # (B-3) 엑셀 저장(10%,20%,30%)
            excel_rows_total = []
            for p in percentages:
                cutoff_idx = int(np.ceil(len(user_spending_total) * p))
                if cutoff_idx <= 0:
                    continue
                actual_v = actual_vals_total[cutoff_idx - 1]
                pred_v = predicted_vals_total[cutoff_idx - 1]
                excel_rows_total.append({
                    "비율": f"{int(p * 100)}%",
                    "실제공급가액": actual_v,  # 컬럼 이름 수정 (억 제거)
                    "예측공급가액": pred_v  # 컬럼 이름 수정 (억 제거)
                })

            df_excel_total = pd.DataFrame(excel_rows_total)

            # 1e8을 곱하여 값 변환
            df_excel_total['실제공급가액'] *= 1e8
            df_excel_total['예측공급가액'] *= 1e8

            # "년도" 컬럼 추가 및 설명 생성
            df_excel_total.insert(0, '년도', '전체')  # "년도" 컬럼 추가
            description = "\n".join(
                f"{row['비율']} - 실제공급가액: {row['실제공급가액']}, 예측공급가액: {row['예측공급가액']}"
                for _, row in df_excel_total.iterrows()
            )
            df_excel_total['설명'] = ""  # "설명" 컬럼 추가
            df_excel_total.loc[0, '설명'] = description  # 첫 번째 행에 설명 추가

            # 엑셀 파일 저장
            excel_path_total = os.path.join(output_dir_xlsx, "연도별_VIP_유저.xlsx")
            df_excel_total_copy_file = df_excel_total.copy()
            try:
                df_excel_total_copy_file.to_excel(excel_path_total, index=False)
                print(f"전체(모든 연도 합산) VIP 유저 예측 데이터 Excel 파일 저장 완료: {excel_path_total}")
            except Exception as e:
                print(f"Excel 파일 저장 중 오류 발생: {e}")
                # (B-3) 그래프 (단일 그래프에 파랑 vs 노랑 + 10%,20%,30% 점선)
            try:
                fig_tot = go.Figure()

                # 파랑 영역
                fig_tot.add_trace(go.Scatter(
                    x=x_vals_total,
                    y=actual_vals_total,
                    fill='tozeroy',
                    mode='none',
                    fillcolor='blue',
                    name='실제(모든연도)'
                ))
                # 노랑 영역
                fig_tot.add_trace(go.Scatter(
                    x=x_vals_total,
                    y=predicted_vals_total,
                    fill='tonexty',
                    mode='none',
                    fillcolor='yellow',
                    name='예측(모든연도)'
                ))

                # 점선: 10%,20%,30% (빨강=실제, 오렌지=예측)
                for p in percentages:
                    cutoff_idx = int(np.ceil(len(user_spending_total) * p))
                    if cutoff_idx <= 0:
                        continue
                    cutoff_x = cutoff_idx / len(user_spending_total)

                    # 실제
                    ya = actual_vals_total[cutoff_idx - 1]
                    fig_tot.add_trace(go.Scatter(
                        x=[cutoff_x, cutoff_x],
                        y=[0, ya],
                        mode='lines',
                        line=dict(color='red', dash='dash'),
                        name=f"{int(p * 100)}% 경계(실제)"
                    ))
                    # 예측
                    yp = predicted_vals_total[cutoff_idx - 1]
                    fig_tot.add_trace(go.Scatter(
                        x=[cutoff_x, cutoff_x],
                        y=[0, yp],
                        mode='lines',
                        line=dict(color='orange', dash='dash'),
                        name=f"{int(p * 100)}% 경계(예측)"
                    ))

                fig_tot.update_layout(
                    title="전체(모든 연도) VIP 유저 (실제 vs 예측) 누적분포",
                    xaxis=dict(title="유저 비율", range=[0, 1]),
                    yaxis=dict(title="누적 금액 (억)", range=[0, max(predicted_vals_total) * 1.05]),
                    font=dict(size=12),
                    legend=dict(orientation="h", y=-0.2),
                    margin=dict(l=50, r=50, t=50, b=100)
                )

                # 그래프 저장
                html_path_tot = os.path.join(output_dir_html, "연도별_VIP_유저.html")
                png_path_tot = os.path.join(output_dir_png, "연도별_VIP_유저.png")
                save_plotly_fig(fig_tot, html_path_tot, png_path_tot)
            except Exception as e:
                print(f"그래프 생성 중 오류 발생: {e}")

        except Exception as e:
            print(f"VIP Users 분석 중 오류 발생: {e}")
    except Exception as e:
        print(f"VIP Users 분석 중 오류 발생: {e}")




# ----------------------------
# Financial Prediction Function
# ----------------------------
def predict_next_year_for_each_year(net_profit):

    net_profit = net_profit.copy()
    net_profit['년도'] = net_profit['년도'].astype(int)
    net_profit['예측매출'] = np.nan
    net_profit['예측판관비'] = np.nan
    net_profit['예측당기순이익'] = np.nan

    years = sorted(net_profit['년도'].unique())
    for y in years:
        # 올해까지의 데이터를 사용하여 내년을 예측
        train_df = net_profit[net_profit['년도'] <= y]
        if train_df.empty:
            continue

        # LinearRegression
        lr = LinearRegression()

        future_year = y + 1
        # x,y
        X = train_df[['년도']].values
        y_sale = train_df['매출'].values
        y_cost = train_df['판관비'].values
        y_profit = train_df['당기순이익'].values

        # 매출 예측
        lr.fit(X, y_sale)
        pred_sale = lr.predict([[future_year]])[0]

        # 판관비 예측
        lr.fit(X, y_cost)
        pred_cost = lr.predict([[future_year]])[0]

        # 당기순이익 예측
        lr.fit(X, y_profit)
        pred_profit = lr.predict([[future_year]])[0]

        # 현재 년도에 다음 년도의 예측치 입력
        net_profit.loc[net_profit['년도'] == y, '예측매출'] = pred_sale
        net_profit.loc[net_profit['년도'] == y, '예측판관비'] = pred_cost
        net_profit.loc[net_profit['년도'] == y, '예측당기순이익'] = pred_profit

    return net_profit


# ----------------------------
# Additional Utility Functions for Area Analysis
# ----------------------------
def load_geojson(geo_file_path):
    """
    Load GeoJSON file.
    """
    try:
        with open(geo_file_path, 'r', encoding='utf-8') as f:
            geo = json.load(f)
        return geo
    except Exception as e:
        print(f"GeoJSON 파일 로딩 중 오류 발생: {e}")
        return None


def map_region_coordinates(geo):
    """
    Map region codes to their first coordinate (or centroid).
    기존 코드는 Polygon/MultiPolygon에 대해 centroid(중심점)를 계산했지만,
    여기서는 "제일 앞에 있는 좌표값"을 가져오도록 수정.
    """
    region_coordinates = {}
    try:
        for feature in geo.get('features', []):
            properties = feature.get('properties', {})
            geometry = feature.get('geometry', {})

            # 지역 코드 추출 (GeoJSON의 실제 필드명에 맞게 수정)
            # [CHANGED] 예: 만약 SIG.geojson에서 'SIG_CD'라면 아래처럼 수정
            # region_code = properties.get('SIG_CD')
            # 일단 원본의 adm_cd / code / id를 그대로 두되, 필요 시 'SIG_CD'로 교체
            region_code = (
                    properties.get('adm_cd')
                    or properties.get('code')
                    or properties.get('id')
                    or properties.get('SIG_CD')  # 필요한 필드명 추가
            )
            if not region_code:
                # 지역 코드가 없을 경우 건너뜀
                continue
            region_code = str(region_code).zfill(5)  # 5자리로 패딩

            # 지역 이름 추출 (필요 시)
            region_name = (
                    properties.get('adm_nm')
                    or properties.get('name')
                    or properties.get('address')
                    or properties.get('SIG_KOR_NM')  # 필요한 필드명 추가
            )
            if not region_name:
                # 지역 이름이 없을 경우 건너뜀
                continue

            # [CHANGED] 좌표 계산 로직을 "제일 앞의 좌표"로 변경
            geom_type = geometry.get('type')
            coords = geometry.get('coordinates', [])

            if geom_type == 'Point':
                # 예: coordinates: [경도, 위도]
                lon, lat = coords[0], coords[1]

            elif geom_type == 'Polygon':
                # 예: Polygon -> 2중 배열 구조 [[ [lon, lat], [lon, lat], ... ]]
                # "제일 앞에 있는 좌표" = coords[0][0]
                # 만약 centroid를 사용하려면 기존처럼 np.mean(...) 사용
                if coords and coords[0]:
                    first_coord = coords[0][0]
                    lon, lat = first_coord[0], first_coord[1]
                else:
                    continue

            elif geom_type == 'MultiPolygon':
                # 예: MultiPolygon -> 3중 배열 구조 [[[ [lon, lat], ... ]], [ ... ]]
                # "제일 앞에 있는 좌표" = coords[0][0][0]
                if coords and coords[0] and coords[0][0]:
                    first_coord = coords[0][0][0]
                    lon, lat = first_coord[0], first_coord[1]
                else:
                    continue

            else:
                # 지원하지 않는 geometry 타입
                continue

            # 좌표 저장 (위도, 경도 순)
            region_coordinates[region_code] = (lat, lon)

    except Exception as e:
        print(f"지역 좌표 매핑 중 오류 발생: {e}")
    return region_coordinates


def save_map_as_png(html_file_path, png_file_path):
    """
    Save a Folium map (HTML) as a PNG file using Selenium.
    """
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1200x900")

    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()),
        options=chrome_options
    )

    try:
        driver.get(f"file://{os.path.abspath(html_file_path)}")
        time.sleep(2)  # Wait for the map to fully render

        driver.save_screenshot(png_file_path)
        print(f"PNG saved at '{png_file_path}'")
    except Exception as e:
        print(f"맵을 PNG로 저장 중 오류 발생: {e}")
    finally:
        driver.quit()


# ----------------------------
# Area Analysis Function
# ----------------------------

def analyze_area(oracle_area, geo_file_path, region_data,
                output_dir_xlsx, output_dir_html, output_dir_png):
    """
    Perform area-wise sales analysis, generate corresponding bubble maps, and save top 5 data to Excel files.
    Additionally, save matched and unmatched region codes to separate TXT files.
    """
    try:
        print("analyze_area 함수가 호출되었습니다.")

        matched_regions = set()
        unmatched_regions = set()
        print("Initialized matched_regions and unmatched_regions sets.")

        print("지역코드 매핑 시작.")
        oracle_area['지역코드'] = oracle_area['지역'].map(region_data)

        print("지역코드 정수형으로 변환 시작.")
        oracle_area['지역코드'] = oracle_area['지역코드'].dropna().astype(int).astype(str).str.zfill(5)
        print(f"매핑된 지역코드: {oracle_area['지역코드'].unique()[:10]}")

        oracle_area['지역코드'] = oracle_area['지역코드'].fillna('NaN')
        print("지역코드 매핑 완료.")

        # 주석 처리된 print 문을 제거하거나 주석으로 남겨둡니다.
        # print(oracle_area.columns)  # 여기 주석 처리 필요

        print("매출 데이터 필터링 및 병합 시작.")
        merged_user_data = oracle_area.copy()
        print(f"매출 데이터 필터링 완료. 병합된 데이터 개수: {len(merged_user_data)}")

        print("GeoJSON 파일 로드 시작.")
        geo = load_geojson(geo_file_path)
        if geo is None:
            print("GeoJSON 파일 로드 실패.")
            return

        print("지역 좌표 매핑 시작.")
        region_coordinates = map_region_coordinates(geo)
        region_coordinates = {str(k).zfill(5): v for k, v in region_coordinates.items()}
        print(f"지역 좌표 매핑 완료: {len(region_coordinates)} 지역.")

        print("유저별 지역 및 연도별 공급가액 집계 시작.")
        merged_user_area = merged_user_data[['지역', '지역코드', '년도', '공급가액']]
        user_supply_sum = merged_user_area.groupby(['지역', '지역코드', '년도'])['공급가액'].sum().reset_index()
        print(f"Aggregated user supply sum:\n{user_supply_sum.head()}")

        # 모든 연도를 확인
        unique_years = sorted(user_supply_sum['년도'].unique())
        print(f"유니크한 년도들: {unique_years}")

        combined_top5_dict = {}
        print("Initialized combined_top5_dict.")

        print("지역코드 매칭 여부 확인 시작.")
        unique_region_codes = oracle_area['지역코드'].unique()
        for code in unique_region_codes:
            if code in region_coordinates:
                matched_regions.add(code)
                print(f"매칭 성공: {code}")
            else:
                if code != 'NaN':
                    unmatched_regions.add(code)
                    print(f"매칭 실패: {code}")
        print("지역코드 매칭 여부 확인 완료.")

        for year in unique_years:
            try:
                print(f"\n년도별 분석 시작: {year}")
                year_data = user_supply_sum[user_supply_sum['년도'] == year]
                print(f"년도 {year}의 데이터 개수: {len(year_data)}")
                print(f"년도 {year}의 데이터 샘플:\n{year_data.head()}")

                year_dir_html = os.path.join(output_dir_html, str(year))
                year_dir_png = os.path.join(output_dir_png, str(year))
                year_dir_xlsx = os.path.join(output_dir_xlsx, str(year))

                os.makedirs(year_dir_html, exist_ok=True)
                os.makedirs(year_dir_png, exist_ok=True)
                os.makedirs(year_dir_xlsx, exist_ok=True)

                print(f"생성된 디렉토리: {year_dir_html}, {year_dir_png}, {year_dir_xlsx}")

                map_center = [35.96, 127.1]
                map_year = folium.Map(location=map_center, zoom_start=7, tiles='cartodbpositron')
                print("Folium 맵 초기화 완료.")

                for _, row in year_data.iterrows():
                    region_code = str(row['지역코드']).zfill(5)
                    supply_value = row['공급가액']
                    if region_code in region_coordinates:
                        lat, lon = region_coordinates[region_code]
                        bubble_size = max(supply_value / 1e6, 1)
                        folium.CircleMarker(
                            location=[lat, lon],
                            radius=bubble_size,
                            fill=True,
                            fill_color='skyblue',
                            fill_opacity=0.6,
                            stroke=False,
                            popup=f'지역 코드: {region_code}<br>공급가액: {supply_value:,.0f}원'
                        ).add_to(map_year)

                        matched_regions.add(region_code)
                    else:

                        unmatched_regions.add(region_code)

                html_file_path = os.path.join(year_dir_html, f'{year}_지역별_판매량.html')
                map_year.save(html_file_path)
                print(f"Folium 맵 HTML 저장 완료: {html_file_path}")

                png_file_path = os.path.join(year_dir_png, f'{year}_지역별_판매량.png')
                try:
                    save_map_as_png(html_file_path, png_file_path)
                    print(f"맵 PNG 저장 완료: {png_file_path}")
                except Exception as e:
                    print(f"맵 PNG 저장 중 오류 발생: {e}")

                top5_year = year_data.sort_values(by='공급가액', ascending=False).head(5)


                # 지역 이름 추가
                detailed_top5 = pd.merge(
                    merged_user_area,
                    top5_year,
                    on=['지역코드', '년도'],
                    suffixes=('', '_total')
                )
                print(f"상위 5개 지역에 대한 상세 데이터:\n{detailed_top5.head()}")

                detailed_top5 = pd.merge(
                    detailed_top5,
                    oracle_area[['지역코드', '지역']].drop_duplicates(),
                    on='지역코드',
                    how='left'
                )

                # '지역_x'와 '지역_y' 컬럼 제거하고 최종적으로 '지역'만 남김
                detailed_top5['지역'] = detailed_top5['지역_y']  # '지역_y'를 '지역'으로 통합
                detailed_top5 = detailed_top5.drop(columns=['지역_x', '지역_y'])

                print(f"'지역' 이름 포함 후 데이터:\n{detailed_top5.head()}")

                top5_year_area = detailed_top5[['지역', '공급가액']].drop_duplicates()
                sum_by_area = top5_year_area.groupby('지역').sum().reset_index().sort_values(by='공급가액', ascending=False)
                print(f"지역별 공급가액 합산:\n{sum_by_area}")

                top5_year_area = sum_by_area.head(5)



                # '년도' 컬럼을 첫 번째 열로 삽입
                top5_year_area_with_year = top5_year_area.copy()
                top5_year_area_with_year.insert(0, '년도', year)

                top5_year_area_with_year['예측 공급가액'] = np.nan  # 초기화

                top5_year_area['예측 공급가액'] = np.nan

                for 지역 in top5_year_area['지역']:
                    지역_data = detailed_top5[detailed_top5['지역'] == 지역][['년도', '공급가액']].drop_duplicates()
                    지역_data = 지역_data.sort_values('년도')

                    if len(지역_data) < 2:
                        # 데이터가 부족하면 증가율 적용
                        print(f"지역 {지역}: 데이터 부족으로 증가율 기반 예측")
                        최근_공급가액 = 지역_data['공급가액'].iloc[-1] if not 지역_data.empty else 0
                        증가율 = 1.1  # 10% 증가
                        top5_year_area.loc[top5_year_area['지역'] == 지역, '예측 공급가액'] = 최근_공급가액 * 증가율
                        continue

                    X = 지역_data['년도'].values.reshape(-1, 1)
                    y = 지역_data['공급가액'].values
                    model = LinearRegression()
                    model.fit(X, y)

                    다음_연도 = 지역_data['년도'].max() + 1
                    예측값 = model.predict([[다음_연도]])[0]
                    top5_year_area.loc[top5_year_area['지역'] == 지역, '예측 공급가액'] = 예측값
                    print(f"지역 {지역}: {다음_연도}년 예측 공급가액 = {예측값}")

                # 결과 저장
                top5_year_area_with_year = top5_year_area.copy()
                top5_year_area_with_year.insert(0, '년도', year)

                # 조건문 추가: 2024년일 경우 예측값 += 공급가액
                if year == 2024:
                    print("2024년 데이터: 예측 공급가액에 공급가액 더하기")
                    top5_year_area_with_year['예측 공급가액'] += top5_year_area_with_year['공급가액']

                # 디렉토리 생성 및 엑셀 저장
                year_dir_xlsx = os.path.join(output_dir_xlsx, str(year))
                os.makedirs(year_dir_xlsx, exist_ok=True)
                # 설명 생성: 각 행에 대한 설명 텍스트 작성
                description = "\n".join(
                    f"{row['년도']}년도 {row['지역']} - 공급가액: {row['공급가액']}, 예측 공급가액: {row['예측 공급가액']}"
                    for _, row in top5_year_area_with_year.iterrows()
                )

                # 설명을 첫 번째 행에 추가하고 나머지 행은 공백으로 설정
                top5_year_area_with_year['설명'] = ""
                top5_year_area_with_year.iloc[0, top5_year_area_with_year.columns.get_loc('설명')] = description

                excel_file_path = os.path.join(year_dir_xlsx, f'{year}_지역별_판매량.xlsx')
                with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
                    top5_year_area_with_year.to_excel(writer, sheet_name='상위5_집계', index=False)

                print(f"{year}년 지역별 판매량 Excel 저장 완료 (예측 포함): {excel_file_path}")


                combined_top5_dict[year] = top5_year_area_with_year.copy()
                combined_top5_dict[year]['년도'] = year
                print(f"combined_top5_dict에 년도 {year}의 데이터 추가 완료.")

            except Exception as e:
                print(f"년도 {year} 처리 중 오류 발생: {e}")

        # 전체 지역별 공급가액 집계
        try:
            user_supply_sum_total = merged_user_area.groupby(['지역코드'])['공급가액'].sum().reset_index()
            print("전체 지역별 공급가액 집계 완료.")
            print(f"전체 지역별 공급가액:\n{user_supply_sum_total.head()}")

            combined_map = folium.Map(location=[35.96, 127.1], zoom_start=7, tiles='cartodbpositron')
            print("전체 지역용 Folium 맵 초기화 완료.")

            for _, row in user_supply_sum_total.iterrows():
                region_code = str(row['지역코드']).zfill(5)
                supply_value = row['공급가액']
                if region_code in region_coordinates:
                    lat, lon = region_coordinates[region_code]
                    bubble_size = max(supply_value / 5e6, 1)
                    folium.CircleMarker(
                        location=[lat, lon],
                        radius=bubble_size,
                        fill=True,
                        fill_color='skyblue',
                        fill_opacity=0.6,
                        stroke=False,
                        popup=f'지역 코드: {region_code}<br>공급가액: {supply_value:,.0f}원'
                    ).add_to(combined_map)
                    matched_regions.add(region_code)
                else:
                    unmatched_regions.add(region_code)

            combined_html_path = os.path.join(output_dir_html, "연도별_지역별_판매량.html")
            combined_map.save(combined_html_path)
            print(f"전체 지역별 Folium 맵 HTML 저장 완료: {combined_html_path}")

            combined_png_path = os.path.join(output_dir_png, "연도별_지역별_판매량.png")
            try:
                save_map_as_png(combined_html_path, combined_png_path)
                print(f"전체 지역별 맵 PNG 저장 완료: {combined_png_path}")
            except Exception as e:
                print(f"전체 맵 PNG 저장 중 오류 발생: {e}")
            # 모든 연도 데이터를 병합


            # 병합된 데이터프레임에서 head() 호출
            combined_top5_df = pd.concat(combined_top5_dict.values(), ignore_index=True)


            combined_top5_df = combined_top5_df.groupby('지역').sum().reset_index()
            combined_top5_df = combined_top5_df[['지역', '공급가액']].sort_values(by='공급가액', ascending=False)
            # 초기화
            combined_top5_df['예측공급가액'] = np.nan  # 초기화
            # 각 지역별로 선형 회귀 적용
            for 지역 in combined_top5_df['지역'].unique():
                # 지역 데이터 필터링
                지역_data = user_supply_sum[user_supply_sum['지역'] == 지역].copy()
                # 데이터프레임 확인
                if not isinstance(지역_data, pd.DataFrame):
                    raise ValueError(f"지역 {지역}의 데이터가 DataFrame이 아닙니다: {type(지역_data)}")

                지역_data = 지역_data.sort_values('년도')  # 연도 정렬

                if len(지역_data) < 2:
                    # 데이터가 2개 미만이면 예측 불가, 최근 공급가액을 그대로 사용
                    최근_공급가액 = 지역_data['공급가액'].iloc[-1] if not 지역_data.empty else 0
                    combined_top5_df.loc[combined_top5_df['지역'] == 지역, '예측공급가액'] = 최근_공급가액
                    continue

                # 선형 회귀 학습
                X = 지역_data['년도'].values.reshape(-1, 1)  # 연도
                y = 지역_data['공급가액'].values  # 공급가액
                model = LinearRegression()
                model.fit(X, y)

                # 다음 연도 예측
                다음_연도 = 지역_data['년도'].max() + 1
                예측값 = model.predict([[다음_연도]])[0]
                combined_top5_df.loc[combined_top5_df['지역'] == 지역, '예측공급가액'] = 예측값

            # 예측값과 공급가액 합산
            combined_top5_df['예측공급가액'] += combined_top5_df['공급가액']

            # '년도' 열을 "전체"로 설정
            combined_top5_df['년도'] = "전체"
            combined_top5_df = combined_top5_df[['년도'] + [col for col in combined_top5_df.columns if col != '년도']]

            # 데이터 정렬 및 상위 5개 선택
            combined_top5_df = combined_top5_df.sort_values(by='공급가액', ascending=False)
            combined_top5_df = combined_top5_df.head(5)

            # 엑셀 저장
            combined_excel_path = os.path.join(output_dir_xlsx, "연도별_지역별_판매량.xlsx")
            combined_top5_df.to_excel(combined_excel_path, index=False)

            # 엑셀 파일 읽기
            combined_data = pd.read_excel(combined_excel_path)

            # 설명 문자열 생성
            description = "\n".join(
                f"{row['지역']} - 공급가액: {row['공급가액']}, 예측 공급가액: {row['예측공급가액']}"
                for _, row in combined_data.iterrows()
            )

            # 설명 추가 (첫 번째 행에만 설명 추가)
            combined_data['설명'] = ""
            combined_data.iloc[0, combined_data.columns.get_loc('설명')] = description

            # 엑셀 파일 다시 저장
            combined_data.to_excel(combined_excel_path, index=False)

        except Exception as e:
            print(f"전체 분석 과정 중 오류 발생: {e}")
    except Exception as e:
        print(f"전체 분석 과정 중 오류 발생: {e}")

def list_xlsx_files(directory):
    path = Path(directory)
    return list(path.rglob('*.xlsx'))

# ----------------------------
# Main Processing Function
# ----------------------------
def process_all_analysis():
    """
    Main function to orchestrate all analysis tasks.
    """
    try:
        # File paths
        input_file = './merged/merged_data.xlsx'
        geo_file_path = './유저/SIG.geojson'
        region_file_path = './유저/region_data.json'

        # Create output directories
        paths = create_output_paths()
        output_dir_xlsx = paths["output_dir_xlsx"]
        output_dir_html = paths["output_dir_html"]
        output_dir_png = paths["output_dir_png"]

        # Ensure output directories exist
        os.makedirs(output_dir_xlsx, exist_ok=True)
        os.makedirs(output_dir_html, exist_ok=True)
        os.makedirs(output_dir_png, exist_ok=True)


        # 기존의 데이터 로드 및 분석 코드
        oracle_data, oracle_item, oracle_area = retrieve_oracle_data()
        merged_data = pd.read_excel(input_file)



        # Perform calculations
        sales_data, sales_by_year = calculate_sales(merged_data)
        cost_by_year = calculate_cost(merged_data)
        net_profit = calculate_net_profit(sales_by_year, cost_by_year)
        net_profit['년도'] = net_profit['년도'].astype(int)

        # Predict future data
        net_profit = predict_next_year_for_each_year(net_profit)

        # Add total row to dataset
        total_row = {
            '년도': '전체',
            '매출': net_profit['매출'].sum(),
            '판관비': net_profit['판관비'].sum(),
            '당기순이익': net_profit['당기순이익'].sum(),
            '예측매출': net_profit['예측매출'].sum(),
            '예측판관비': net_profit['예측판관비'].sum(),
            '예측당기순이익': net_profit['예측당기순이익'].sum(),
        }
        total_row_df = pd.DataFrame([total_row])
        combined_net_profit = total_row_df

        # 각 연도별 그래프 (올해 vs 내년)
        min_year = net_profit['년도'].min()
        max_year = net_profit['년도'].max()
        for y in range(min_year, max_year + 1):
            # 해당 연도의 실제값
            row_h = net_profit[net_profit['년도'] == y]
            if row_h.empty:
                continue
            # dict화
            hist_data = {
                '매출': row_h['매출'].values[0],
                '판관비': row_h['판관비'].values[0],
                '당기순이익': row_h['당기순이익'].values[0]
            }
            # 올해 행에 "예측"은 (내년 예측)
            pred_data = {
                '매출': row_h['예측매출'].values[0],
                '판관비': row_h['예측판관비'].values[0],
                '당기순이익': row_h['예측당기순이익'].values[0]
            }

            # 각 연도별 그래프 생성
            plot_year_with_prediction(y, hist_data, pred_data, output_dir_html, output_dir_png)

        # 전체 그래프 생성
        plot_full_prediction_with_actuals(net_profit, output_dir_html, output_dir_png)

        # Save overall financial data to Excel
        financial_output_path = os.path.join(output_dir_xlsx, "연도별_재무지표.xlsx")

        # 설명 생성: 전체 데이터를 요약한 텍스트 추가
        description = "\n".join(
            f"년도: {row['년도']} - 매출: {row['매출']}, 판관비: {row['판관비']}, 당기순이익: {row['당기순이익']}, "
            f"예측 매출: {row['예측매출']}, 예측 판관비: {row['예측판관비']}, 예측 당기순이익: {row['예측당기순이익']}"
            for _, row in combined_net_profit.iterrows()
        )

        combined_net_profit['설명'] = ""
        combined_net_profit.loc[0, '설명'] = description  # 첫 번째 행에 설명 추가

        combined_net_profit.to_excel(financial_output_path, index=False)
        print(f"전체 데이터를 포함한 Excel 저장 완료: {financial_output_path}")

        # Save yearly financial data
        for year in net_profit['년도'].unique():
            year_dir = os.path.join(output_dir_xlsx, str(year))
            os.makedirs(year_dir, exist_ok=True)
            year_data = net_profit[net_profit['년도'] == year]

            # 설명 생성: 해당 연도의 데이터를 요약한 텍스트 추가
            year_description = "\n".join(
                f"년도: {row['년도']} - 매출: {row['매출']}, 판관비: {row['판관비']}, 당기순이익: {row['당기순이익']}, "
                f"예측 매출: {row['예측매출']}, 예측 판관비: {row['예측판관비']}, 예측 당기순이익: {row['예측당기순이익']}"
                for _, row in year_data.iterrows()
            )
            year_data['설명'] = ""
            year_data.iloc[0, year_data.columns.get_loc('설명')] = year_description

            year_file_path = os.path.join(year_dir, f"{year}_재무지표.xlsx")
            year_data.to_excel(year_file_path, index=False)
            print(f"{year}년 재무지표 저장 완료: {year_file_path}")

        # Perform various analyses
        analyze_category(merged_data, oracle_item, output_dir_xlsx, output_dir_html, output_dir_png)
        analyze_gender(merged_data, oracle_data, output_dir_xlsx, output_dir_html, output_dir_png)
        analyze_age_group(merged_data, oracle_data, output_dir_xlsx, output_dir_html, output_dir_png)
        analyze_vip_users(merged_data, oracle_data, output_dir_xlsx, output_dir_html, output_dir_png)

        # Check required files
        for path in [input_file, geo_file_path, region_file_path]:
            if not os.path.exists(path):
                raise FileNotFoundError(f"Required file not found: {path}")

        # Load region data
        with open(region_file_path, "r", encoding="utf-8") as f:
            region_data = json.load(f)

        # Perform area analysis
        analyze_area(
            oracle_area, geo_file_path,
            region_data, output_dir_xlsx, output_dir_html, output_dir_png
        )
        # 1) 최종 결과를 저장할 새 Workbook(엑셀) 생성
        master_wb = Workbook()
        master_ws = master_wb.active
        master_ws.title = "MergedData"  # 시트 이름(원하시면 변경 가능)

        # 2) 폴더 안의 모든 .xlsx 파일 목록 가져오기
        xlsx_files = list(Path(output_dir_xlsx).rglob("*.xlsx"))

        # 3) 파일 하나씩 열어서, 행 단위로 마스터 시트에 append
        for file_path in xlsx_files:
            # 예: 내부 관리용 파일(master_data.xlsx 등)은 건너뛰고 싶다면 다음과 같이 처리
            # if file_path.name == "master_data.xlsx":
            #     continue

            print(f"[INFO] '{file_path.name}' 파일 처리 중...")

            # (a) 원본 파일 열기
            wb = load_workbook(file_path)

            # (b) 특정 시트만 쓴다면 여기에서 sheet 이름 지정 가능
            #     기본적으로 첫 번째 시트를 사용한다고 가정
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]

            # (c) 원본 파일의 모든 행을 순회하며, 최종 마스터 시트에 그대로 append
            for row in ws.iter_rows(values_only=True):
                master_ws.append(row)

        # 4) 최종 워크북을 파일로 저장
        output_file_path = os.path.join(output_dir_xlsx, "엑셀데이터종합본.xlsx")
        master_wb.save(output_file_path)
        print(f"[DONE] 모든 파일을 이어 붙인 결과가 '{output_dir_xlsx}' 에 저장되었습니다.")


    except FileNotFoundError as e:
        print(e)
    return True, "모든 분석 작업이 완료되었습니다."